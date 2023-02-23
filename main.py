'''
genetic algorithm for a scheduling problem

2/22~2/26 產生40條可行解

'''
import math
import time
import random
import pandas as pd
import openpyxl
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
pd.options.mode.chained_assignment = None

#換線表轉換為字典
def transfor_lineTable(lineTable_path):
    table = pd.read_excel(lineTable_path,skiprows=2)
    table = table.drop(table.columns[0],axis=1).set_index(table.columns[1])
    return table.to_dict()

#可生產對照表轉換為字典
def transfor_manuTable(manuTable_path):
    table = pd.read_excel(manuTable_path)
    table = table.drop(table.columns[0],axis=1).set_index(table.columns[1])
    return table.to_dict()

#在order上標記類型
def orderLabel(order,typeTable):
    order['類型'] = None
    for i in order.index:
        for j in typeTable.index:
            if order['產品品號'][i]==typeTable['途程品號'][j]:
                order['類型'][i] = typeTable['類別'][j]
                break
    return order

#解碼成各自的排程表陣列
def easyDecode(individual):
    '''
    將染色體切回各自排程表陣列
    例如
    [x1,2,3,4,x2,5,6,x3,9,7]
    =>
    [
        [2,3,4]
        [5,6]
        [9,7]
    ]

    做法:
    宣告tmp，每次遇到字串就表示接下來都屬於下一個陣列，把tmp加入schedule並清空
    最後還要在append一次tmp

    '''

    schedule = []
    tmp = []

    for e in individual:
        if type(e)==str:

            #tmp裡面有值才要append到schedule
            #如果tmp非空就會回傳True
            if tmp:
                schedule.append(tmp)
                tmp = []
        else:
            tmp.append(e)

    schedule.append(tmp)


    # print(individual)
    # for lst in schedule:
    #     for e in lst:
    #         print(e,end=' ')
    #     print()

    return schedule

#將當前時間換算回日期
def convertDate(dailySheet,cate,value):
    date = ''
    # print(f"\n{cate}班 當前時間: {value}")

    row = 0
    while value>0:
        minutes = dailySheet[str(cate)+'班'][row]

        # print(f"日期={dailySheet['日期'][row]}\
        #     value={value} 當日剩餘工時={minutes}")

        value -= minutes
        if value>0:
            row += 1
        else:
            value = -(value)
            break

    date = dailySheet['日期'][row]

    # print(f"轉換為日期: {date} 當日剩餘工時: {value}")
    return date

#產生每日工時表
def generate_dailySheet(fillTable_path,LOST):

    #讀取待填工時表的新增工時欄位，並加上加班時數
    #數值存放於dailySheet，一開始為二維陣列，最後轉換成dataFrame return
    #範例格式:  [
    #           [日期1,1班工時,2班工時,3班工時],
    #           [日期2,1班工時,2班工時,3班工時],
    #           [日期3,1班工時,2班工時,3班工時]
    #       ]

    dailySheet = []

    wb_fillTable = load_workbook(fillTable_path)
    ws = wb_fillTable['待填工時表']

    #最大工班數，從第六欄開始，每個製造工班佔三個欄位
    maxCate = (ws.max_column-5)//3

    #取得新增工時，並加上加班時間(乘上損耗率)
    for row in range(3,ws.max_row+1):
        date = ws['A'+str(row)].value

        #跳過空白列
        if date is None:continue

        tmp = [date]
        stdTimeList = []
        if type(ws['E'+str(row)].value) == int:
            record = 0
            ws['E'+str(row)].value = str(ws['E'+str(row)].value)
            for i in range(1,len(ws['E'+str(row)].value)+1):
                if i == len(ws['E'+str(row)].value):break
                if i%3==0:
                    stdTimeList.append(ws['E'+str(row)].value[record:i])
                    record = i
            stdTimeList.append(ws['E'+str(row)].value[record:])
        else:
            stdTimeList = ws['E'+str(row)].value.split(',')

        #查看當天各班是否有加班
        index = 8
        for cate in range(maxCate):
            overTime = ws[get_column_letter(index)+str(row)].value

            if overTime is None:
                overTime = 0
            else:
                overTime = round(overTime * 60 * (1-float(LOST[cate])/100))

            index += 3
            tmp.append(int(stdTimeList[cate])+overTime)

        dailySheet.append(tmp)

    # dailySheet轉換為dataframe

    #新增欄位
    col = ['日期']
    for cate in range(maxCate):
        col.append(str(cate+1)+'班')

    dailySheet = pd.DataFrame(dailySheet,columns=col)

    # dailySheet.to_excel('./dailySheet.xlsx')
    # quit()

    return dailySheet

#建立crnTable
def create_crntTable(dailySheet,ws_fillTable):
    crnTable = pd.DataFrame()
    crnTable = pd.DataFrame(columns=['工班','開始工作時間','目前加工類型'])
    crnTable['工班'] = dailySheet.drop('日期',axis=1).columns
    crnTable.set_index('工班',inplace=True)

    #第6欄為製造一班的剩餘工時
    x=6
    cateIdx = 1
    for i in crnTable.index:
        crnTable['開始工作時間'][i] = 0

        #讀取「前一筆訂單類型」
        crnTable['目前加工類型'][i] = ws_fillTable[get_column_letter(x)+'1'].value

        #將補正還有剩餘工時加到開始工作時間
        #如果是空白則給0
        corretHour = ws_fillTable[get_column_letter(x)+'3'].value
        if corretHour is None:
            corretHour = 0

        remaningHour = ws_fillTable[get_column_letter(x+1)+'3'].value
        if remaningHour is None:
            remaningHour = 0

        startWorkTime = corretHour + remaningHour

        crnTable['開始工作時間'][str(cateIdx)+'班'] = startWorkTime

        cateIdx += 1

        #每個製造工班有3個欄位，換下個工班時就+3
        x += 3

    return crnTable

def manufHours(order,typeTable_path,count):
    #根據人數計算加工時間，並寫入order
    #捆綁訂單的加工時間需要保留,因此先分離兩部分，等加工時間全部算完再連接
    # bundle = ['特殊類型','其餘特殊']

    # bundleOrder = order.copy()
    # for i in order.index:
    #     if order['類型'][i] in bundle:
    #         order.drop(i,inplace = True)
    #     else:
    #         bundleOrder.drop(i,inplace = True)

    # order.to_excel('input/debug/order.xlsx')
    # bundleOrder.to_excel('input/debug/bun.xlsx')

    #新增欄位 台/分，並且重置為None
    order['台/分'] = None
    order['加工時間'] = None

    #讀取機種對照表
    typeTable = load_workbook(typeTable_path)
    ws = typeTable['產品途程明細表 (主檔) 20190214']


    #根據人數選擇縱軸欄位
    #固定欄位:
    # 11 : 'E'
    # 8  : 'G'
    # 6  : 'H'

    col = ''
    if count==11:
        col = 'E'
    elif count==8:
        col = 'G'
    elif count==6:
        col = 'H'


    #迴圈寫入對應類型加工時間
    for i in order.index:
        for row in range(3,ws.max_row):
            if order['產品品號'][i] == ws['A'+str(row)].value:

                #讀取對應單位工時
                hour = ws[col + str(row)].value
                if type(hour)==str:
                    continue

                # 加工時間(分) = 產量 / (一分鐘做幾個)
                time = math.ceil(order['產量'][i] / hour)
                order['台/分'][i] = hour
                order['加工時間'][i] = time
                break

    #連接捆綁訂單
    # order = pd.concat([order,bundleOrder])


    order.reset_index(drop=True,inplace=True)


    return order

#計算開工與完工
def calculating_done(cate,df,lineTable,dailySheet,fillTable_path):

    #讀取待填工時表
    ws_fillTable = load_workbook(fillTable_path)['待填工時表']


    #初始化crnTable
    crnTable = pd.DataFrame()
    crnTable = create_crntTable(dailySheet,ws_fillTable)



    #更新排程表的換線時間欄位
    type = crnTable['目前加工類型'][cate]

    for i in df.index:
        if df['類型'][i]!=type:
            type = df['類型'][i]


    #更新預計開工與完工
    for i in df.index:

        #當前時間加上換線時間
        lineTime = 0
        type = crnTable['目前加工類型'][cate]
        if df['類型'][i] != type:
            lineTime = lineTable[df['類型'][i]][type]

        crnTable['開始工作時間'][cate] += lineTime


        #當前時間加上 "加工時間" 與更新當前類型
        ori = convertDate(dailySheet,cate[:1],crnTable['開始工作時間'][cate])

        crnTable['開始工作時間'][cate] += df['加工時間'][i]
        crnTable['目前加工類型'][cate] = df['類型'][i]

        #完工日期存入crn
        crn = convertDate(dailySheet,cate[:1],crnTable['開始工作時間'][cate])
        df['預計開工'][i] = ori
        df['預計完工'][i] = crn


    return df

#可行解檢查
def check_feasibility(individual,order,manuTable,lineTable,dailySheet,fillTable_path):
    '''
    (a) 該工班能不能做這張單
    (b) 做完會不會超過交期


    (a) test1 :
        1.簡單解碼後去逐個比對，只要發現一筆違反限制就break

    (b) test2:
        1.將完整製令單依照各班排程表切分成df並放入df_schedule
        2.傳入calculating_done 計算完工時間
        3.確認是否所有訂單都沒有逾期 (完工日>出貨日)，一樣只要發現一筆就break

    '''
    result = True
    test1 = True
    test2 = True
    schedule = easyDecode(individual)
    df_schedule = []

    crew_num = 1
    for lst in schedule:
        if not test1:break

        for idx in lst:
            type = order['類型'][idx]
            if manuTable[f'製{crew_num}班11人'][type]==0:
                test1 = False
                break
        crew_num += 1

    # print(schedule)
    # print(test1)
    # quit()

    #計算各班訂單完工時間
    crew_num = 1
    for new_index in schedule:
        df = pd.DataFrame()
        df = order.reindex(new_index)
        df = calculating_done(str(crew_num)+'班',df,lineTable,dailySheet,fillTable_path)

        df.to_excel(f'./debug/{crew_num}.xlsx')

        df_schedule.append(df)
        crew_num += 1

    #檢查是否有超過交期
    crew_num = 1
    for df in df_schedule:
        if not test2:break
        for idx in df.index:
            complete = pd.to_datetime(df['預計完工'][idx], format='%Y/%m/%d')

            due = df['預計出貨'][idx]
            if complete > due:
                # print(str(crew_num)+'班',idx,complete,due)
                test2 = False
                # quit()
                break
        crew_num += 1

    result = test1 and test2

    return result

#初始化染色體
def init_individual(order,manuTable):
    '''
    產生染色體:
    1.讀取製令單index並打亂
    2.依照亂數分配給工班
    3.檢查工班是否可做 不可以->重新產生一個亂數
        a.因為這邊沒有計算加工時間，所以人數都用11人去看
        b.只要那張訂單分配到的工班不能做就會重複產生亂數
    4.組合成染色體的編碼格式

    '''
    individual = []
    X1 = ['X1']
    X2 = ['X2']
    X3 = ['X3']
    X4 = ['X4']
    order_list = order.index.tolist()
    random.shuffle(order_list)


    for idx in order_list:
        type = order['類型'][idx]
        while(True):
            rnd = round(random.uniform(0,1),2)
            if rnd<=0.25:
                if manuTable['製1班11人'][type]==1:
                    X1.append(idx)
                    break
                else:continue

            elif rnd<=0.5:
                if manuTable['製2班11人'][type]==1:
                    X2.append(idx)
                    break
                else:continue

            elif rnd<=0.75:
                if manuTable['製3班11人'][type]==1:
                    X3.append(idx)
                    break
                else:continue

            elif rnd<=1:
                if manuTable['製4班11人'][type]==1:
                    X4.append(idx)
                    break
                else:continue

    for lst in [X1,X2,X3,X4]:
        for element in lst:
            individual.append(element)

    # print(X1)
    # print(X2)
    # print(X3)
    # print(X4)
    # print(individual)
    # quit()

    return individual

#評估適應度
def fitness(population):
    return

#選擇
def selection(population):
    return

#交配
def crossover(mating_pool):
    return

#突變
def mutation(population):
    return

#主程式區塊
def main():
    POPULATION_SIZE = 40
    MAX_GENERATION = 10
    MUTATION_RATE = 0.2
    LOST = [3,3,3,3]
    population = []

    #資料預處理
    print(f'[{round(time.process_time(),2)}s] 資料預處理')

    fillTable_path = './待填工時表單-20221121-2-2-2-3.xlsx'
    typeTable_path = './福佑電機製造部工時總攬資料_V3.xlsx'
    dailySheet = generate_dailySheet(fillTable_path,LOST)

    order = pd.read_excel('./製令單_1121-1125.xlsx')
    typeTable = pd.read_excel(typeTable_path,skiprows=1)
    lineTable = transfor_lineTable('./換線表測試V3.xlsx')
    manuTable = transfor_manuTable('./工時及可生產產品對應_V2.xlsx')

    order = orderLabel(order,typeTable)
    order = manufHours(order,typeTable_path,11)

    # print('訂單筆數:',len(order))


    #初始化族群
    print(f'[{round(time.process_time(),2)}s] 初始化族群')

    individual_count = 1
    for i in range(POPULATION_SIZE):

        individual = init_individual(order,manuTable)
        while(check_feasibility(individual,order,manuTable,lineTable,dailySheet,fillTable_path)!=True):
            individual = init_individual(order,manuTable)

        print(f'[{round(time.process_time(),2)}s] 可行解數目:{individual_count}')
        individual_count += 1
        population.append(individual)

    # for individual in population:
    #     print(individual)


    #演化
    for generation in range(MAX_GENERATION):

        best_fitness = fitness(population)

        mating_pool = selection(population)

        offspring = crossover(mating_pool)

        population = offspring

        population = mutation(population)

    print(f'[{round(time.process_time(),2)}s] done')

T = 1
for i in range(5):
    print(f'\n[{round(time.process_time(),2)}s] 執行週期: {T}')
    main()
    T += 1
