import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
import math
import datetime
import openpyxl
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font,PatternFill,Alignment,Border,Side
import os

# 換線表轉換為字典
def transfor_lineTable(lineTable_path):
    table = pd.read_excel(lineTable_path,skiprows=2)
    table = table.drop(table.columns[0],axis=1).set_index(table.columns[1])
    return table.to_dict()

# 可生產對照表轉換為字典
def transfor_manuTable(manuTable_path):
    table = pd.read_excel(manuTable_path)
    table = table.drop(table.columns[0],axis=1).set_index(table.columns[1])
    return table.to_dict()

# 在order上標記類型
def orderLabel(order,typeTable):
    order['類型'] = None
    for i in order.index:
        for j in typeTable.index:
            if order['產品品號'][i]==typeTable['途程品號'][j]:
                order['類型'][i] = typeTable['類別'][j]
                break
    return order

# 將當前時間換算回日期
def convertDate(dailySheet,cate,value):
    date = ''
    # print(f"\n{cate}班 當前時間: {value}")

    row = 0
    while value>0:
        minutes = dailySheet[str(cate)+'班'][row]

        # print(f"日期={dailySheet['日期'][row]}\
        #     value={value} 當日剩餘工時={minutes}")

        value -= minutes
        if value>0:
            row += 1
        else:
            value = -(value)
            break

    date = dailySheet['日期'][row]

    # print(f"轉換為日期: {date} 當日剩餘工時: {value}")
    return date

# 產生每日工時表
def generate_dailySheet(fillTable_path,LOST):

    #讀取待填工時表的新增工時欄位，並加上加班時數
    #數值存放於dailySheet，一開始為二維陣列，最後轉換成dataFrame return
    #範例格式:  [
    #           [日期1,1班工時,2班工時,3班工時],
    #           [日期2,1班工時,2班工時,3班工時],
    #           [日期3,1班工時,2班工時,3班工時]
    #       ]

    dailySheet = []

    wb_fillTable = load_workbook(fillTable_path)
    ws = wb_fillTable['待填工時表']

    #最大工班數，從第六欄開始，每個製造工班佔三個欄位
    maxCate = (ws.max_column-5)//3

    #取得新增工時，並加上加班時間(乘上損耗率)
    for row in range(3,ws.max_row+1):
        date = ws['A'+str(row)].value

        #跳過空白列
        if date is None:continue

        tmp = [date]
        stdTimeList = []
        if type(ws['E'+str(row)].value) == int:
            record = 0
            ws['E'+str(row)].value = str(ws['E'+str(row)].value)
            for i in range(1,len(ws['E'+str(row)].value)+1):
                if i == len(ws['E'+str(row)].value):break
                if i%3==0:
                    stdTimeList.append(ws['E'+str(row)].value[record:i])
                    record = i
            stdTimeList.append(ws['E'+str(row)].value[record:])
        else:
            stdTimeList = ws['E'+str(row)].value.split(',')

        #查看當天各班是否有加班
        index = 8
        for cate in range(maxCate):
            overTime = ws[get_column_letter(index)+str(row)].value

            if overTime is None:
                overTime = 0
            else:
                overTime = round(overTime * 60 * (1-float(LOST[cate])/100))

            index += 3
            tmp.append(int(stdTimeList[cate])+overTime)

        dailySheet.append(tmp)

    # dailySheet轉換為dataframe

    #新增欄位
    col = ['日期']
    for cate in range(maxCate):
        col.append(str(cate+1)+'班')

    dailySheet = pd.DataFrame(dailySheet,columns=col)

    # dailySheet.to_excel('./dailySheet.xlsx')
    # quit()

    return dailySheet

# 建立crnTable
def create_crntTable(dailySheet,ws_fillTable):
    crnTable = pd.DataFrame()
    crnTable = pd.DataFrame(columns=['工班','開始工作時間','目前加工類型'])
    crnTable['工班'] = dailySheet.drop('日期',axis=1).columns
    crnTable.set_index('工班',inplace=True)

    #第6欄為製造一班的剩餘工時
    x=6
    cateIdx = 1
    for i in crnTable.index:
        crnTable['開始工作時間'][i] = 0

        #讀取「前一筆訂單類型」
        crnTable['目前加工類型'][i] = ws_fillTable[get_column_letter(x)+'1'].value

        #將補正還有剩餘工時加到開始工作時間
        #如果是空白則給0
        corretHour = ws_fillTable[get_column_letter(x)+'3'].value
        if corretHour is None:
            corretHour = 0

        remaningHour = ws_fillTable[get_column_letter(x+1)+'3'].value
        if remaningHour is None:
            remaningHour = 0

        startWorkTime = corretHour + remaningHour

        crnTable['開始工作時間'][str(cateIdx)+'班'] = startWorkTime

        cateIdx += 1

        #每個製造工班有3個欄位，換下個工班時就+3
        x += 3

    return crnTable

# 計算加工時間
def manufHours(order,typeTable_path,count):
    #根據人數計算加工時間，並寫入order
    #捆綁訂單的加工時間需要保留,因此先分離兩部分，等加工時間全部算完再連接
    # bundle = ['特殊類型','其餘特殊']

    # bundleOrder = order.copy()
    # for i in order.index:
    #     if order['類型'][i] in bundle:
    #         order.drop(i,inplace = True)
    #     else:
    #         bundleOrder.drop(i,inplace = True)

    # order.to_excel('input/debug/order.xlsx')
    # bundleOrder.to_excel('input/debug/bun.xlsx')

    #新增欄位 台/分，並且重置為None
    order['台/分'] = None
    order['加工時間'] = None

    #讀取機種對照表
    typeTable = load_workbook(typeTable_path)
    ws = typeTable['產品途程明細表 (主檔) 20190214']


    #根據人數選擇縱軸欄位
    #固定欄位:
    # 11 : 'E'
    # 8  : 'G'
    # 6  : 'H'

    col = ''
    if count==11:
        col = 'E'
    elif count==8:
        col = 'G'
    elif count==6:
        col = 'H'


    #迴圈寫入對應類型加工時間
    for i in order.index:
        for row in range(3,ws.max_row):
            if order['產品品號'][i] == ws['A'+str(row)].value:

                #讀取對應單位工時
                hour = ws[col + str(row)].value
                if type(hour)==str:
                    continue

                # 加工時間(分) = 產量 / (一分鐘做幾個)
                time = math.ceil(order['產量'][i] / hour)
                order['台/分'][i] = hour
                order['加工時間'][i] = time
                break

    #連接捆綁訂單
    # order = pd.concat([order,bundleOrder])


    order.reset_index(drop=True,inplace=True)


    return order

# 計算開工與完工
def calculating_done(cate,df,lineTable,dailySheet,fillTable_path):

    #讀取待填工時表
    ws_fillTable = load_workbook(fillTable_path)['待填工時表']


    #初始化crnTable
    crnTable = pd.DataFrame()
    crnTable = create_crntTable(dailySheet,ws_fillTable)



    #更新排程表的換線時間欄位
    type = crnTable['目前加工類型'][cate]

    for i in df.index:
        if df['類型'][i]!=type:
            type = df['類型'][i]


    #更新預計開工與完工
    for i in df.index:

        #當前時間加上換線時間
        lineTime = 0
        type = crnTable['目前加工類型'][cate]
        if df['類型'][i] != type:
            lineTime = lineTable[df['類型'][i]][type]

        crnTable['開始工作時間'][cate] += lineTime


        #當前時間加上 "加工時間" 與更新當前類型
        ori = convertDate(dailySheet,cate[:1],crnTable['開始工作時間'][cate])

        crnTable['開始工作時間'][cate] += df['加工時間'][i]
        crnTable['目前加工類型'][cate] = df['類型'][i]

        #完工日期存入crn
        crn = convertDate(dailySheet,cate[:1],crnTable['開始工作時間'][cate])
        df['預計開工'][i] = ori
        df['預計完工'][i] = crn


    return df


# 將排程表按照各周分開
def splitWeek(dfList):
    # 使用方式:
    # table[指定工班][週數]
    # ex: table['2班'][2] 將可取得2班第二週的排程表

    #更改排程表顯示欄位
    for i in range(len(dfList)):
        dfList[i] = dfList[i][[
                '預計開工',
                '預計完工',
                '預計出貨',
                '產品品號',
                '品名',
                '規格',
                '製令編號',
                '產量',
                '類型',
                '台/分',
                '加工時間',
                '換線時間',
                '選班'
                ]]

    col = []
    tmp = []
    for i in range(len(dfList)):
        col.append(str(i+1)+'班')
        tmp.append(None)
    table = pd.DataFrame(columns=col)

    #直接初始化五個禮拜
    for i in range(5):
        table.loc[len(table)] = tmp


    for df in dfList:
        cate = df['選班'][df.index[0]]
        count=0

        #設定week欄位
        df['week'] = None
        for i in df.index:
            df['week'][i] = datetime.datetime.strptime(df['預計開工'][i],'%Y/%m/%d').isocalendar()[1]
        
        #建立週別的陣列
        allWeek = list(dict.fromkeys(df['week'].tolist()))

        #將對應範圍的排程表放入table
        for week in allWeek:
            table[cate][count] = df[df['week']==week].drop('week',axis=1)
            table[cate][count].reset_index(inplace=True,drop=True)
            count += 1


    table.index = range(1,len(table)+1)

    return table

#建立輸出排程表
def createSheet(fillTable_path,dailySheet,weekTable,lost):

    today = datetime.datetime.today().strftime("%Y%m%d")
    wb_schedule = openpyxl.Workbook()
    wb_fill = load_workbook(fillTable_path)
    ws_fill = wb_fill['待填工時表']


    #初始化所有sheet
    #如果第n週有任何1班的數值不為None，再建立sheet
    fix = 8
    for week in weekTable.index:
        for cate in weekTable.columns:
            if weekTable[cate][week] is not None:
                #從待填工時表去抓當週一跟六 當成sheetName
                end = ws_fill['A'+str(fix)].value
                fix += 7
                start = datetime.datetime.strptime(end,'%Y/%m/%d')-\
                    datetime.timedelta(days=5)
                start = datetime.datetime.strftime(start,'%Y-%m-%d')
                end = end.replace('/','-')
                sheetName = f"{start}~{end}"
                wb_schedule.create_sheet(sheetName)
                break

    del wb_schedule['Sheet']
    wb_schedule.save(f"output/排程表測試_{today}.xlsx")

    #新增excel欄位
    week = 0
    fix = 8

    #remaining 紀錄(本)剩餘工時
    tmp = []
    for i in weekTable.columns:tmp.append(0)
    remaining = pd.Series(tmp,index=weekTable.columns)

    for sheetName in wb_schedule.sheetnames:
        week += 1


        ws_schedule = wb_schedule[sheetName]

        #從待填工時表去抓當週一跟六
        end = ws_fill['A'+str(fix)].value
        fix += 7
        start = datetime.datetime.strptime(end,'%Y/%m/%d')-\
        datetime.timedelta(days=5)
        start = datetime.datetime.strftime(start,'%Y/%m/%d')


        firstTable = True
        for cate in weekTable.columns:
            if weekTable[cate][week] is None:
                continue

            #用write 控制各班排程表間的空白列
            if firstTable:
                write = ws_schedule.max_row-1
                firstTable = False
            else:
                write = ws_schedule.max_row+1

            #在這裡就讀取排程表，欄位要填入時會用到
            df = weekTable[cate][week]


            #新增"製造n班"欄位
            ws_schedule.merge_cells(f"A{str(write+1)}:C{str(write+2)}")
            ws_schedule['A'+str(write+1)] = '製造' + cate

            #新增參考欄位

            #稼動總工時: 一週可用工時+加班  (需考慮損耗率
            ws_schedule['A'+str(write+3)] = '稼動總工時:'
            ws_schedule['B'+str(write+3)] = dailySheet.set_index('日期')[cate][start:end].sum()


            #(上)剩餘工時: 第一週讀待填工時表，第二週開始為上一週的(本)剩餘工時
            ws_schedule['A'+str(write+4)] = '(上)剩餘工時:'
            if week==1:
                x = int(cate[0])*3+3
                value = ws_fill[get_column_letter(x)+'3'].value

                if value is None:
                    ws_schedule['B'+str(write+4)] = 0
                else:
                    ws_schedule['B'+str(write+4)] = value
            else:
                ws_schedule['B'+str(write+4)] = remaining[cate]

            #(本)新增工時: 本週所有工單的總加工時間
            ws_schedule['A'+str(write+5)] = '(本)新增工時:'
            ws_schedule['B'+str(write+5)] = df['加工時間'].sum()

            #
            ws_schedule['C'+str(write+3)] = '損耗率(%):'
            ws_schedule['D'+str(write+3)] = str(float(lost[int(cate[0])-1]))+'%'

            #換線補正: 本週排程表換線時間總和
            ws_schedule['C'+str(write+4)] = '換線補正:'
            ws_schedule['D'+str(write+4)] = df['換線時間'].sum()


            #(補正工時): 第一週讀待填工時表，第二週開始為0
            ws_schedule['E'+str(write+4)] = '補正工時:'

            if week==1:
                x = int(cate[0])*3+4
                value = ws_fill[get_column_letter(x)+'3'].value
                ws_schedule['F'+str(write+4)] = value
                if value is None:
                    ws_schedule['F'+str(write+4)] = 0
                else:
                    ws_schedule['F'+str(write+4)] = value
            else:
                ws_schedule['F'+str(write+4)] = 0


            #(本)剩餘工時: 稼動總工時 – (本)新增工時 – 換線補正 – 補正工時
            ws_schedule['C'+str(write+5)] = '(本)剩餘工時:'
            ws_schedule['D'+str(write+5)] = \
                ws_schedule['B'+str(write+3)].value\
                -ws_schedule['B'+str(write+5)].value\
                -ws_schedule['D'+str(write+4)].value\
                -ws_schedule['F'+str(write+4)].value
            remaining[cate] = ws_schedule['D'+str(write+5)].value * -1
            if remaining[cate]<0:
                ws_schedule['D'+str(write+5)].fill = PatternFill(start_color='FFE66F', fill_type="solid")

            #加班工時(小時): 把這週加班小時都加起來
            ws_schedule['E'+str(write+5)] = '加班工時:'
            #x: 該工班對應行
            #y: 根據week找出該週第一天，用來控制while迴圈
            #z: 當週對應列為 z~z+5
            x = int(cate[0])*3+5
            y = 1
            z = 3
            while week!=y:
                y += 1
                z += 7
            sum = 0
            for i in range(z,z+6):
                tmp = ws_fill[get_column_letter(x)+str(i)].value
                if tmp is not None:
                    sum += tmp

            ws_schedule['F'+str(write+5)] = sum


            #
            ws_schedule['G'+str(write+3)] = '製令別:'
            ws_schedule['H'+str(write+3)] = '510廠內製令'

            ws_schedule['G'+str(write+4)] = '製表日期:'
            ws_schedule['H'+str(write+4)] = datetime.datetime.today().strftime("%Y/%m/%d")

            ws_schedule['G'+str(write+5)] = '備註:'
            ws_schedule['H'+str(write+5)] = ''

            #附加排程表
            df['預計出貨'] = df['預計出貨'].astype("string")
            for i in df.index:
                df['預計出貨'][i] = df['預計出貨'][i].replace('-','/')
                writer = pd.ExcelWriter(f"output/排程表測試_{today}.xlsx", engine='openpyxl')
                writer.book = wb_schedule
                writer.sheets = dict((ws.title, ws) for ws in wb_schedule.worksheets)
                df.drop('選班',axis=1).to_excel(writer,startrow=write+6,sheet_name=sheetName,index=False)
            writer.save()

            #把逾期的訂單hightlight
            for row in range(write+8,ws_schedule.max_row+1):

                end_time = datetime.datetime.strptime(ws_schedule['B'+str(row)].value,'%Y/%m/%d')
                due_time = datetime.datetime.strptime(ws_schedule['C'+str(row)].value,'%Y/%m/%d')
                
                if end_time>=due_time:
                    for col in range(1,ws_schedule.max_column+1):
                        char=get_column_letter(col)
                        ws_schedule[char+str(row)].fill = PatternFill(start_color='FFE66F', fill_type="solid")



            ws = wb_schedule[sheetName]
            #縮放
            ws.sheet_view.zoomScale=135

            #欄寬
            for col in range(1,11):
                ws.column_dimensions[get_column_letter(col)].width=16
            ws.column_dimensions["F"].width=27

            for row in range(1,ws.max_row+1):
                for col in range(1,ws.max_column+1):
                    char=get_column_letter(col)
                    cell=ws[char + str(row)]

                    #文字對齊
                    cell.alignment=Alignment(horizontal="left")

            #製造x班
            ws_schedule['A'+str(write+1)].font=Font(size=24)


    wb_schedule.save(f"output/排程表測試_{today}.xlsx")


#移動該列到最上面
def move_to_top(df,index):
    #新增new欄位並依序填入1~len+1
    df["new"] = range(1,len(df)+1)

    #搜尋到想移動的那列並將new更改為0
    df.loc[df.index==index, 'new'] = 0

    #以new欄排序 此時剛剛那列就會到最上方，再移除new欄
    df = df.sort_values("new").drop('new', axis=1)
    return df

#移動該列到最下面
def move_to_bottom(df,index):
    #新增new欄位並依序填入1~len+1
    df["new"] = range(1,len(df)+1)

    #搜尋到想移動的那列並將new更改為len+2
    df.loc[df.index==index, 'new'] = len(df)+2

    #以new欄排序 此時剛剛那列就會到最下方，再移除new欄
    df = df.sort_values("new").drop('new', axis=1)

    return df

#換線順序調整
def dfimization(lineTable,df,crt,next):

    #0. 拆出兩日工單
    df.drop(crt.index,inplace=True)
    df.drop(next.index,inplace=True)
    crt_tag_1 = crt[crt['tag']==1]
    crt_tag_0 = crt[crt['tag']==0]

    #1. 對今日tag_0及下一天排序
    crt_tag_0 = crt_tag_0.sort_values('類型')
    crt = pd.concat([crt_tag_1,crt_tag_0])
    next = next.sort_values('類型')


    #2. 從tag_0中尋找今明最多類型  #! 先不管tag 直接全部一起判斷 評估哪種做法比較好
    # crtType = list(dict.fromkeys(crt_tag_0['類型'].tolist()))
    crtType = list(dict.fromkeys(crt['類型'].tolist()))

    nextType = next['類型'].tolist()
    most = None
    max = 0

    for t in crtType:
        if nextType.count(t)>max:
            max = nextType.count(t)
            most = t

    #都沒有相同類型就找最小換線時間
    if most is None:
        nextType = list(dict.fromkeys(nextType))

        minTime = 999
        minTypeCrt = None
        minTypeNext = None

        for i in crtType:
            for j in nextType:
                tmp = lineTable[i][j]
                # print(i,j,tmp)
                if tmp<minTime:
                    minTime = tmp
                    minTypeCrt = i
                    minTypeNext = j

        # print(minTime,minTypeCrt,minTypeNext)



    #3. 移動最多類型或最小換線類型並標記tag為1
    if most is not None:
        for row in crt.index:
            if crt['類型'][row]==most and crt['tag'][row]==0:
                crt = move_to_bottom(crt,row)

        for row in next.index:
            if next['類型'][row]==most:
                next=move_to_top(next,row)
                next['tag'][row]=1
    else:
        for row in crt.index:
            if crt['類型'][row]==minTypeCrt and crt['tag'][row]==0:
                crt = move_to_bottom(crt,row)

        for row in next.index:
            if next['類型'][row]==minTypeNext:
                next=move_to_top(next,row)
                next['tag'][row]=1

    #3. 移動最多類型並標記tag為1
    if most!=None:
        for row in crt.index:
            if crt['類型'][row]==most and crt['tag'][row]==0:

                crt = move_to_bottom(crt,row)

        for row in next.index:
            if next['類型'][row]==most:
                next=move_to_top(next,row)
                next['tag'][row]=1



    #4. 合併回df

    #先將這兩天合併並且重新給index
    tmp = pd.concat([crt,next])
    tmp.index = range(tmp.index.min(),tmp.index.max()+1,1)
    #連接原df後再以index排序，讓它排到正確的索引
    df = pd.concat([tmp,df])
    df = df.sort_index()


    # print('test:',crt['日期'][crt.index[0]],'、',next['日期'][next.index[0]])
    # for i in crt.index:
    #     print(i,crt['製令編號'][i],crt['日期'][i],'\t',crt['類型'][i],'\t',crt['tag'][i],)
    # print()
    # for i in next.index:
    #     print(i,next['製令編號'][i],next['日期'][i],'\t',next['類型'][i],'\t',next['tag'][i],)
    # print()



    return df


